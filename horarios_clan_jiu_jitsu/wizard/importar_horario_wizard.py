# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import datetime
import csv
import logging
import re
from datetime import timedelta

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from odf import opendocument, table
except ImportError:
    opendocument = None

class GymHorarioImportarWizard(models.TransientModel):
    _name = 'gym.horario.importar.wizard'
    _description = 'Asistente para Importar Horario – Traductor Universal Robusto'

    excel_file = fields.Binary(string='Archivo (XLSX, XLS, ODS, CSV)', required=True)
    file_name = fields.Char(string='Nombre del Archivo')
    area = fields.Selection([
        ('1', 'AREA 1'),
        ('2', 'AREA 2'),
        ('3', 'AREA 3'),
        ('4', 'AREA 4'),
    ], string='Área de Importación', required=True, default='1')

    def action_import(self):
        if not self.file_name:
            raise UserError(_('No se pudo detectar el nombre del archivo.'))
        
        ext = self.file_name.lower().split('.')[-1]
        file_data = base64.b64decode(self.excel_file)
        
        if ext == 'xlsx':
            data_blocks = self._process_xlsx(file_data)
        elif ext == 'xls':
            data_blocks = self._process_xls(file_data)
        elif ext == 'ods':
            data_blocks = self._process_ods(file_data)
        elif ext == 'csv':
            data_blocks = self._process_csv(file_data)
        else:
            raise UserError(_('Formato "%s" no soportado.') % ext)

        if not data_blocks:
            raise UserError(_('No se encontraron horarios válidos.'))

        self.env['gym.horario'].search([('area', '=', self.area)]).unlink()
        self._create_annual_schedule(data_blocks)
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('¡Importación Exitosa!'),
                'message': _('Se detectaron %s bloques de clase.') % len(data_blocks),
                'sticky': False,
            }
        }

    # --- XLSX Logic ---
    def _process_xlsx(self, file_data):
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            sheet = workbook.active
            processed_merged = []
            blocks = []
            
            time_col = 1
            for r in range(1, 15):
                if self._parse_time(sheet.cell(row=r, column=1).value):
                    time_col = 1; break
                if self._parse_time(sheet.cell(row=r, column=2).value):
                    time_col = 2; break

            for row in range(1, sheet.max_row + 1):
                time_val = sheet.cell(row=row, column=time_col).value
                hora_inicio_row = self._parse_time(time_val)
                if hora_inicio_row is None: continue

                for col in range(time_col + 1, time_col + 8):
                    if col > sheet.max_column: break
                    cell = sheet.cell(row=row, column=col)
                    val = str(cell.value or "").strip()
                    if not val or 'libre' in val.lower(): continue
                    
                    cell_id = f"{row}-{col}"
                    if cell_id in processed_merged: continue

                    max_row = row
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            max_row = merged_range.max_row
                            for r in range(merged_range.min_row, merged_range.max_row + 1):
                                for c in range(merged_range.min_col, merged_range.max_col + 1):
                                    processed_merged.append(f"{r}-{c}")
                            break
                    
                    h_start, h_end = self._extract_times_from_text(val)
                    final_start = h_start if h_start is not None else hora_inicio_row
                    final_end = h_end if h_end is not None else min(hora_inicio_row + (max_row - row + 1) * 0.5, 23.99)
                    
                    # Hardening: Garantizar que fin > inicio para evitar error de validación
                    if final_end <= final_start:
                        final_end = final_start + 1.0

                    blocks.append({
                        'day_offset': col - (time_col + 1),
                        'hora_inicio': final_start,
                        'hora_fin': min(final_end, 23.99),
                        'content': val
                    })
            return blocks
        except Exception as e:
            raise UserError(_('Error en XLSX: %s') % str(e))

    # --- XLS (Legacy) Logic ---
    def _process_xls(self, file_data):
        try:
            workbook = xlrd.open_workbook(file_contents=file_data, formatting_info=True)
            sheet = workbook.sheet_by_index(0)
            blocks = []
            processed_cells = set()

            time_col = 0
            for r in range(0, 15):
                if r >= sheet.nrows: break
                if self._parse_time(sheet.cell_value(r, 0)):
                    time_col = 0; break
                if self._parse_time(sheet.cell_value(r, 1)):
                    time_col = 1; break

            for row in range(0, sheet.nrows):
                time_val = sheet.cell_value(row, time_col)
                hora_inicio_row = self._parse_time(time_val)
                if hora_inicio_row is None: continue

                for col in range(time_col + 1, time_col + 8):
                    if col >= sheet.ncols: break
                    if (row, col) in processed_cells: continue
                    val = str(sheet.cell_value(row, col) or "").strip()
                    if not val or 'libre' in val.lower(): continue

                    max_row = row
                    for crange in sheet.merged_cells:
                        rlo, rhi, clo, chi = crange
                        if rlo <= row < rhi and clo <= col < chi:
                            max_row = rhi - 1
                            for r in range(rlo, rhi):
                                for c in range(clo, chi):
                                    processed_cells.add((r, c))
                            break
                    
                    h_start, h_end = self._extract_times_from_text(val)
                    final_start = h_start if h_start is not None else hora_inicio_row
                    final_end = h_end if h_end is not None else min(hora_inicio_row + (max_row - row + 1) * 0.5, 23.99)
                    
                    if final_end <= final_start:
                        final_end = final_start + 1.0

                    blocks.append({
                        'day_offset': col - (time_col + 1),
                        'hora_inicio': final_start,
                        'hora_fin': min(final_end, 23.99),
                        'content': val
                    })
            return blocks
        except Exception as e:
            raise UserError(_('Error en XLS: %s') % str(e))

    # --- ODS Logic (Most robust) ---
    def _process_ods(self, file_data):
        if not opendocument:
            raise UserError(_('odfpy no instalada.'))
        try:
            doc = opendocument.load(io.BytesIO(file_data))
            sheet = doc.spreadsheet.getElementsByType(table.Table)[0]
            raw_rows = sheet.getElementsByType(table.TableRow)
            blocks = []
            
            expanded_rows = []
            for rr in raw_rows:
                repeat = int(rr.getAttribute('numberrowsrepeated') or 1)
                for _ in range(repeat): expanded_rows.append(rr)

            covered = {} 
            MAX_COLS = 50

            # Auto-detección de la columna de tiempo (Ancla)
            time_col = 0
            for r_check in range(min(len(expanded_rows), 15)):
                row_check = expanded_rows[r_check]
                cells_check = row_check.getElementsByType(table.TableCell)
                # Solo revisamos las primeras 3 columnas para el ancla
                for c_idx in range(3):
                    if c_idx < len(cells_check):
                        val = self._get_ods_cell_value(cells_check[c_idx])
                        if self._parse_time(val):
                            time_col = c_idx
                            break
                else: continue
                break

            for r_idx, row in enumerate(expanded_rows):
                cells = row.getElementsByType(table.TableCell)
                logical_row = [None] * MAX_COLS
                
                c_ptr = 0
                for cell in cells:
                    while covered.get(c_ptr, -1) >= r_idx:
                        c_ptr += 1
                    
                    rep_cols = int(cell.getAttribute('numbercolumnsrepeated') or 1)
                    r_span = int(cell.getAttribute('numberrowsspanned') or 1)
                    c_span = int(cell.getAttribute('numbercolumnsspanned') or 1)

                    for r in range(rep_cols):
                        curr_col = c_ptr + r
                        if curr_col < MAX_COLS:
                            logical_row[curr_col] = cell
                            if r_span > 1:
                                covered[curr_col] = r_idx + r_span - 1
                    
                    c_ptr += (rep_cols + c_span - 1)

                time_val = self._get_ods_cell_value(logical_row[time_col])
                hora_inicio_row = self._parse_time(time_val)
                if hora_inicio_row is None: continue

                for c_idx in range(time_col + 1, time_col + 8):
                    cell = logical_row[c_idx]
                    if not cell: continue
                    val = self._get_ods_cell_value(cell).strip()
                    if not val or 'libre' in val.lower(): continue

                    cell_key = (id(cell), c_idx)
                    if any(b.get('cell_key') == cell_key for b in blocks): continue

                    h_start, h_end = self._extract_times_from_text(val)
                    r_span = int(cell.getAttribute('numberrowsspanned') or 1)
                    
                    final_start = h_start if h_start is not None else hora_inicio_row
                    final_end = h_end if h_end is not None else min(hora_inicio_row + r_span * 0.5, 23.99)
                    
                    if final_end <= final_start:
                        final_end = final_start + 1.0

                    blocks.append({
                        'day_offset': c_idx - (time_col + 1),
                        'hora_inicio': final_start,
                        'hora_fin': min(final_end, 23.99),
                        'content': val,
                        'cell_key': cell_key
                    })
            return blocks
        except Exception as e:
            raise UserError(_('Error en ODS: %s') % str(e))

    # --- CSV Logic ---
    def _process_csv(self, file_data):
        try:
            content = file_data.decode('utf-8-sig')
            dialect = csv.Sniffer().sniff(content[:1024])
            reader = csv.reader(io.StringIO(content), dialect)
            rows = list(reader)
            blocks = []
            for r_idx, row in enumerate(rows):
                if not row: continue
                hora_inicio_row = self._parse_time(row[0])
                col_start = 1
                if hora_inicio_row is None and len(row) > 1:
                    hora_inicio_row = self._parse_time(row[1])
                    col_start = 2
                if hora_inicio_row is None: continue

                for col in range(col_start, min(len(row), col_start + 7)):
                    val = row[col].strip()
                    if not val or 'libre' in val.lower(): continue
                    h_start, h_end = self._extract_times_from_text(val)
                    
                    final_start = h_start if h_start is not None else hora_inicio_row
                    final_end = h_end if h_end is not None else (hora_inicio_row + 1.0)
                    
                    if final_end <= final_start:
                        final_end = final_start + 1.0

                    blocks.append({
                        'day_offset': col - col_start,
                        'hora_inicio': final_start,
                        'hora_fin': min(final_end, 23.99),
                        'content': val
                    })
            return blocks
        except Exception as e:
            raise UserError(_('Error en CSV: %s') % str(e))

    # --- Utils ---
    def _extract_times_from_text(self, text):
        try:
            times = re.findall(r'(\d{1,2}:\d{2})\s*(am|pm)', text.lower())
            if len(times) >= 1:
                h_start = self._parse_time(f"{times[0][0]}{times[0][1]}")
                h_end = None
                if len(times) >= 2:
                    h_end = self._parse_time(f"{times[1][0]}{times[1][1]}")
                return h_start, h_end
        except: pass
        return None, None

    def _get_ods_cell_value(self, cell):
        if not cell: return ""
        from odf.text import P
        p_elements = cell.getElementsByType(P)
        if p_elements:
            return "\n".join([str(p) for p in p_elements])
        return cell.getAttribute('value') or ""

    def _create_annual_schedule(self, blocks):
        start_date = datetime.date(2026, 1, 5)
        for block in blocks:
            disc, inst_name = self._parse_content(block['content'])
            instructor = self._get_or_create_instructor(inst_name, disc)
            for week in range(52):
                actual_date = start_date + timedelta(weeks=week, days=block['day_offset'])
                overlap = self.env['gym.horario'].search([
                    ('area', '=', self.area), ('fecha', '=', actual_date),
                    ('hora_inicio', '<', block['hora_fin']), ('hora_fin', '>', block['hora_inicio']),
                ], limit=1)
                if not overlap:
                    self.env['gym.horario'].create({
                        'instructor_id': instructor.id, 'fecha': actual_date,
                        'hora_inicio': block['hora_inicio'], 'hora_fin': block['hora_fin'],
                        'area': self.area, 'state': 'programado',
                    })

    def _parse_time(self, val):
        if not val: return None
        if isinstance(val, (datetime.time, float)):
            if isinstance(val, float) and val < 1:
                total_minutes = val * 24 * 60
                return int(total_minutes // 60) + (int(total_minutes % 60) / 60.0)
            return val.hour + val.minute / 60.0 if hasattr(val, 'hour') else val
        s = str(val).lower().replace(' ', '').strip()
        if not s: return None
        try:
            meridiem = 'am' if 'am' in s else ('pm' if 'pm' in s else None)
            clean_time = s.replace('am', '').replace('pm', '')
            parts = clean_time.split(':')
            if not parts[0].isdigit(): return None
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
            if meridiem == 'pm' and h < 12: h += 12
            elif meridiem == 'am' and h == 12: h = 0
            if 0 <= h <= 23: return h + m / 60.0
        except: pass
        return None

    def _parse_content(self, text):
        text = text.strip()
        lines = text.split('\n')
        clean_lines = []
        for line in lines:
            if not re.search(r'\d{1,2}:\d{2}', line.lower()):
                clean_lines.append(line.strip())
        
        if len(clean_lines) >= 2:
            disc = clean_lines[0].replace('(', '').replace(')', '').strip()
            name = clean_lines[1].replace('(', '').replace(')', '').strip()
            return disc, name
        elif len(clean_lines) == 1:
            line = clean_lines[0]
            parts = line.split(' ')
            if len(parts) > 1:
                return " ".join(parts[:-1]), parts[-1].replace('(', '').replace(')', '').strip()
            return line, "SENSEI"
        return "CLASE", "SENSEI"

    def _get_or_create_instructor(self, name, spec):
        clean_name = name.strip()
        noise_keywords = ['adelante', 'libre', 'instrucciones', 'horario', 'am', 'pm', 'instructor']
        is_valide_name = True
        if not clean_name or len(clean_name) < 2: is_valide_name = False
        if any(kw in clean_name.lower() for kw in noise_keywords): is_valide_name = False
        if re.search(r'\d{1,2}:\d{2}', clean_name): is_valide_name = False
        
        target_name = clean_name if is_valide_name else "SENSEI"
        
        instr = self.env['gym.instructor'].search([('name', '=ilike', target_name)], limit=1)
        if not instr:
            instr = self.env['gym.instructor'].create({
                'name': target_name, 
                'especialidad': spec if is_valide_name else 'DOJO'
            })
        return instr
